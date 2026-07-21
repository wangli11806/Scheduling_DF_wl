"""
客服排班系统 - 后端服务
Flask + SQLite，单文件部署
启动: pip install flask openpyxl && python app.py
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(sys.executable), 'Lib', 'site-packages'))

import sqlite3, json, hashlib, hmac
from datetime import datetime, date
from flask import Flask, request, jsonify, g, send_file, session, redirect
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import calendar

try:
    from chinese_calendar import is_workday
except ImportError:
    is_workday = None

app = Flask(__name__, static_folder=".", static_url_path="")
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # 开发阶段禁用静态文件缓存
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schedule.db")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ==================== 鉴权 ====================

AUTH_FILE = os.path.join(BASE_DIR, "auth_config.json")
AUTH_EXEMPT = ["/login", "/api/auth/login", "/api/auth/logout", "/api/bot/schedules"]


def load_auth_config():
    if not os.path.exists(AUTH_FILE):
        raise RuntimeError(f"auth_config.json 不存在，请先运行: python setup_password.py")
    with open(AUTH_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


_auth_config = load_auth_config()
app.secret_key = _auth_config["secret_key"]
app.config["SESSION_COOKIE_PATH"] = "/"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


def check_credentials(username, password):
    cfg = load_auth_config()
    if username != cfg.get("username"):
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), cfg["salt"].encode(), 200000)
    return hmac.compare_digest(dk.hex(), cfg["password_hash"])


@app.before_request
def require_auth():
    if any(request.path.startswith(p) for p in AUTH_EXEMPT):
        return None
    if not session.get("logged_in"):
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "未登录"}), 401
        next_url = request.full_path.lstrip("/")
        if "://" in next_url or next_url.lstrip("/").startswith("/"):
            next_url = ""
        return redirect("/login" + ("?next=" + next_url if next_url else ""))


# ==================== 鉴权路由 ====================

@app.route("/login")
def login_page():
    if session.get("logged_in"):
        return redirect("/")
    return app.send_static_file("login.html")


# ==================== 页面英文路由 ====================

@app.route("/overview")
def overview_page():
    return app.send_static_file("总览.html")


@app.route("/schedule")
def schedule_page():
    return app.send_static_file("排班表.html")


@app.route("/schedule-mobile")
def schedule_mobile_page():
    return app.send_static_file("排班表_移动端.html")


@app.route("/monthly-schedule")
def monthly_schedule_page():
    return app.send_static_file("月度排班.html")


@app.route("/raw-schedule")
def raw_schedule_page():
    return app.send_static_file("原始排班.html")


@app.route("/work-arrangement")
def work_arrangement_page():
    return app.send_static_file("工作安排.html")


@app.route("/schedule-adjustment")
def schedule_adjustment_page():
    return app.send_static_file("排班调整.html")


@app.route("/employee")
def employee_page():
    return app.send_static_file("员工管理.html")


@app.route("/shift")
def shift_page():
    return app.send_static_file("班次设置.html")


@app.route("/monthly-hours")
def monthly_hours_page():
    return app.send_static_file("月度时长统计.html")


@app.route("/work-stats")
def work_stats_page():
    return app.send_static_file("工作安排统计.html")


# ==================== API 路由 ====================

@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not username:
        return jsonify({"ok": False, "error": "请输入用户名"})
    if not password:
        return jsonify({"ok": False, "error": "请输入密码"})
    if not check_credentials(username, password):
        return jsonify({"ok": False, "error": "用户名或密码错误"})
    session.clear()
    session["logged_in"] = True
    session["username"] = username
    return jsonify({"ok": True})


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/bot/schedules")
def api_bot_schedules():
    token = request.args.get("token", "")
    cfg = load_auth_config()
    if not token or token != cfg.get("bot_token", ""):
        return jsonify({"ok": False, "error": "token 无效"}), 403

    date_str = request.args.get("date", "")
    employee = request.args.get("employee", "").strip()

    db = get_db()

    # 查询当日请假人员
    leave_emps = set(r["employee_name"] for r in db.execute(
        "SELECT DISTINCT employee_name FROM leave_records WHERE type='leave' AND leave_date=?",
        [date_str]
    ).fetchall())

    # 查询该日期所有放休+换休记录，汇总每个员工的放休总时长
    rest_rows = db.execute("""
        SELECT employee_name, SUM(hours) AS rest_hours
        FROM leave_records
        WHERE type IN ('rest','换休') AND leave_date = ?
        GROUP BY employee_name
    """, [date_str]).fetchall()
    rest_map = {r["employee_name"]: r["rest_hours"] or 0 for r in rest_rows}

    if employee:
        rows = db.execute("""
            SELECT s.schedule_date, s.employee_name, s.shift_name,
                   e.team, sh.start_time, sh.end_time, sh.work_hours
            FROM schedules s
            LEFT JOIN employees e ON e.name = s.employee_name
            LEFT JOIN shifts sh ON sh.name = s.shift_name
            WHERE s.schedule_date = ? AND s.employee_name = ? AND e.status = 'active'
        """, [date_str, employee]).fetchall()
    else:
        rows = db.execute("""
            SELECT s.schedule_date, s.employee_name, s.shift_name,
                   e.team, sh.start_time, sh.end_time, sh.work_hours
            FROM schedules s
            LEFT JOIN employees e ON e.name = s.employee_name
            LEFT JOIN shifts sh ON sh.name = s.shift_name
            WHERE s.schedule_date = ? AND e.status = 'active'
            ORDER BY e.team, s.employee_name
        """, [date_str]).fetchall()

    # 查询当日加班/换班记录，汇总每个员工的加班工时
    overtime_rows = db.execute("""
        SELECT employee_name, SUM(hours) AS overtime_hours
        FROM leave_records
        WHERE type IN ('overtime','换班') AND leave_date = ?
        GROUP BY employee_name
    """, [date_str]).fetchall()
    overtime_map = {r["employee_name"]: r["overtime_hours"] or 0 for r in overtime_rows}

    # 查询当日所有换班记录（换班的人当天实际在上班，用于补充shift信息）
    swap_in_rows = db.execute("""
        SELECT lr.employee_name, lr.start_time, lr.end_time, lr.hours, e.team
        FROM leave_records lr
        LEFT JOIN employees e ON e.name = lr.employee_name
        WHERE lr.type IN ('换班','overtime') AND lr.leave_date = ?
    """, [date_str]).fetchall()
    swap_in_map = {}
    for r in swap_in_rows:
        if r["employee_name"] not in swap_in_map:
            swap_in_map[r["employee_name"]] = r

    schedule_emps = set()
    schedules = []
    for r in rows:
        schedule_emps.add(r["employee_name"])
        wh = (r["work_hours"] or 0) + overtime_map.get(r["employee_name"], 0)
        rh = rest_map.get(r["employee_name"], 0)
        if r["employee_name"] in leave_emps:
            working = False
        else:
            working = (wh - rh) > 0
        schedules.append({
            "employee": r["employee_name"],
            "team": r["team"] or "",
            "shift": r["shift_name"],
            "start_time": r["start_time"] or "",
            "end_time": r["end_time"] or "",
            "working": working,
            "rest_hours": round(rh, 1)
        })

    # 仅有加班/换班记录（原排班表无记录或为休息）的员工也加入上班列表
    for emp_name, info in swap_in_map.items():
        if emp_name not in schedule_emps:
            rh = rest_map.get(emp_name, 0)
            wh = info["hours"] or 0
            schedules.append({
                "employee": emp_name,
                "team": info["team"] or "",
                "shift": "加班/换班",
                "start_time": info["start_time"] or "",
                "end_time": info["end_time"] or "",
                "working": (wh - rh) > 0,
                "rest_hours": round(rh, 1)
            })

    return jsonify({
        "ok": True,
        "date": date_str,
        "count": len(schedules),
        "schedules": schedules
    })


def count_working_days(year, month):
    """返回指定月份的中国法定工作日天数（考虑节假日和调休）"""
    days_in_month = calendar.monthrange(year, month)[1]
    if is_workday is None:
        # 降级方案：按周一至周五计算（不含中国节假日）
        return sum(1 for d in range(1, days_in_month + 1)
                   if date(year, month, d).weekday() < 5)
    return sum(1 for d in range(1, days_in_month + 1)
               if is_workday(date(year, month, d)))


# ==================== 日志管理 ====================

class _Tee:
    """同时写入文件和原始流"""
    def __init__(self, *streams):
        self._streams = streams

    def write(self, data):
        for s in self._streams:
            s.write(data)
            s.flush()

    def flush(self):
        for s in self._streams:
            s.flush()

    def __getattr__(self, name):
        return getattr(self._streams[0], name)


def _rotate_log(path):
    """启动时清除旧日志内容"""
    try:
        open(path, "w").close()
    except OSError:
        pass


def setup_logging():
    """将 stdout/stderr 重定向到 logs/ 子目录，同时保留控制台输出"""
    log_dir = os.path.join(BASE_DIR, "logs")
    out_dir = os.path.join(log_dir, "output")
    err_dir = os.path.join(log_dir, "error")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(err_dir, exist_ok=True)

    out_path = os.path.join(out_dir, "service_output.log")
    err_path = os.path.join(err_dir, "service_error.log")

    _rotate_log(out_path)
    _rotate_log(err_path)

    sys.stdout = _Tee(sys.__stdout__, open(out_path, "a", encoding="utf-8"))
    sys.stderr = _Tee(sys.__stderr__, open(err_path, "a", encoding="utf-8"))


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db:
        db.close()


def init_db():
    db = sqlite3.connect(DATABASE)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS employees (
            id          TEXT PRIMARY KEY,
            name        TEXT NOT NULL UNIQUE,
            team        TEXT NOT NULL,
            position    TEXT NOT NULL DEFAULT '客服',
            supervisor     TEXT DEFAULT '',
            dongfu_id      TEXT DEFAULT '',
            work_hour_system TEXT DEFAULT '',
            entry_date     TEXT NOT NULL,
            status      TEXT NOT NULL DEFAULT 'active',
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS shifts (
            id          TEXT PRIMARY KEY,
            name        TEXT NOT NULL UNIQUE,
            info        TEXT DEFAULT '',
            start_time  TEXT DEFAULT '',
            end_time    TEXT DEFAULT '',
            lunch_start TEXT DEFAULT '',
            lunch_end   TEXT DEFAULT '',
            dinner_start TEXT DEFAULT '',
            dinner_end  TEXT DEFAULT '',
            work_hours  REAL DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS schedules (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_date  TEXT NOT NULL,
            employee_name  TEXT NOT NULL,
            shift_name     TEXT NOT NULL,
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(schedule_date, employee_name)
        );
        CREATE INDEX IF NOT EXISTS idx_schedules_date ON schedules(schedule_date);
        CREATE INDEX IF NOT EXISTS idx_schedules_emp  ON schedules(employee_name);

        CREATE TABLE IF NOT EXISTS raw_schedules (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_date  TEXT NOT NULL,
            employee_name  TEXT NOT NULL,
            shift_name     TEXT NOT NULL,
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(schedule_date, employee_name)
        );
        CREATE INDEX IF NOT EXISTS idx_raw_schedules_date ON raw_schedules(schedule_date);
        CREATE INDEX IF NOT EXISTS idx_raw_schedules_emp  ON raw_schedules(employee_name);

        CREATE TABLE IF NOT EXISTS monthly_schedules (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_date  TEXT NOT NULL,
            employee_name  TEXT NOT NULL,
            shift_name     TEXT NOT NULL,
            finalized      INTEGER DEFAULT 0,
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            updated_at     TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(schedule_date, employee_name)
        );
        CREATE INDEX IF NOT EXISTS idx_monthly_schedules_date ON monthly_schedules(schedule_date);

        CREATE TABLE IF NOT EXISTS daily_assignments (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            date           TEXT NOT NULL,
            employee_name  TEXT NOT NULL,
            work_types     TEXT NOT NULL DEFAULT '[]',
            lunch_slot     TEXT DEFAULT '',
            dinner_slot    TEXT DEFAULT '',
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            updated_at     TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(date, employee_name)
        );
        CREATE INDEX IF NOT EXISTS idx_assignments_date ON daily_assignments(date);

        CREATE TABLE IF NOT EXISTS leave_records (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            type           TEXT DEFAULT 'overtime',
            team           TEXT NOT NULL,
            leave_date     TEXT NOT NULL,
            dongfu_id      TEXT DEFAULT '',
            employee_name  TEXT NOT NULL,
            start_time     TEXT DEFAULT '',
            end_time       TEXT DEFAULT '',
            hours          REAL DEFAULT 0,
            remark         TEXT DEFAULT '',
            submitter      TEXT DEFAULT '',
            deduction      REAL DEFAULT 0,
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            updated_at     TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_leave_records_date ON leave_records(leave_date);
    """)
    # 兼容旧表：新增 deduction 列
    try:
        db.execute("ALTER TABLE leave_records ADD COLUMN deduction REAL DEFAULT 0")
    except:
        pass
    # 兼容旧表：新增 type 列
    try:
        db.execute("ALTER TABLE leave_records ADD COLUMN type TEXT DEFAULT 'overtime'")
    except:
        pass
    db.executescript("""
        CREATE TABLE IF NOT EXISTS swap_records (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            swap_type      TEXT NOT NULL,
            person_a       TEXT NOT NULL,
            team_a         TEXT DEFAULT '',
            date_a         TEXT NOT NULL,
            shift_a        TEXT DEFAULT '',
            person_b       TEXT NOT NULL,
            team_b         TEXT DEFAULT '',
            date_b         TEXT NOT NULL,
            shift_b        TEXT DEFAULT '',
            remark         TEXT DEFAULT '',
            operator       TEXT DEFAULT '',
            created_at     TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_swap_records_date_a ON swap_records(date_a);
        CREATE INDEX IF NOT EXISTS idx_swap_records_date_b ON swap_records(date_b);
        CREATE INDEX IF NOT EXISTS idx_leave_records_team ON leave_records(team);
    """)
    db.commit()

    # 迁移：为已有数据库添加 dongfu_id 列
    try:
        db.execute("ALTER TABLE employees ADD COLUMN dongfu_id TEXT DEFAULT ''")
        db.commit()
    except sqlite3.OperationalError:
        pass

    # 迁移：为已有数据库添加 work_hour_system 列
    try:
        db.execute("ALTER TABLE employees ADD COLUMN work_hour_system TEXT DEFAULT ''")
        db.commit()
    except sqlite3.OperationalError:
        pass

    # 写入默认数据
    cur = db.execute("SELECT COUNT(*) FROM employees")
    if cur.fetchone()[0] == 0:
        db.executescript("""
            INSERT INTO employees(id, name, team, position, supervisor, entry_date) VALUES
                ('EMP001','张三','在线组','客服','陈敏','2024-03-15'),
                ('EMP002','李四','热线组','客服','陈敏','2024-06-01'),
                ('EMP003','王芳','售后组','客服','陈敏','2024-08-20'),
                ('EMP004','赵磊','在线组','客服','陈敏','2025-01-10'),
                ('EMP005','陈敏','主管',  '主管','',    '2023-05-08');
        """)
    cur = db.execute("SELECT COUNT(*) FROM shifts")
    if cur.fetchone()[0] == 0:
        db.executescript("""
            INSERT INTO shifts(id, name, info, start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end, work_hours) VALUES
                ('SHF001','A班','早1班',  '08:00','17:00','12:00','12:30','','',8.0),
                ('SHF002','B班','常规班','09:00','18:00','12:00','12:30','','',8.5),
                ('SHF003','C班','早2班',  '07:00','16:00','12:00','12:30','','',8.5),
                ('SHF004','D班','晚1班',  '14:00','23:00','','','18:00','18:30',8.5),
                ('SHF005','E班','通宵班','22:00','08:00','','','02:00','02:30',9.5);
        """)
    db.commit()
    db.close()


# 导入时自动初始化数据库（兼容所有启动方式）
init_db()


# ==================== 员工 API ====================

@app.route("/api/employees", methods=["GET"])
def api_employees_list():
    db = get_db()
    status = request.args.get("status", "").strip()
    if status in ("active", "inactive"):
        rows = db.execute("SELECT * FROM employees WHERE status=? ORDER BY id", (status,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM employees ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/employees", methods=["POST"])
def api_employees_create():
    data = request.json
    name = (data.get("name") or "").strip()
    team = (data.get("team") or "").strip()
    position = (data.get("position") or "").strip()
    supervisor = (data.get("supervisor") or "").strip()
    entry_date = (data.get("entryDate") or "").strip()
    dongfu_id = (data.get("dongfuId") or "").strip()
    work_hour_system = (data.get("workHourSystem") or "").strip()
    status = (data.get("status") or "active").strip()

    if not name:
        return jsonify({"error": "请输入员工姓名"}), 400
    if not team:
        return jsonify({"error": "请选择所属团队"}), 400
    if not position:
        return jsonify({"error": "请选择岗位"}), 400

    db = get_db()
    exist = db.execute("SELECT id FROM employees WHERE name=?", (name,)).fetchone()
    if exist:
        return jsonify({"error": f"员工\"{name}\"已存在"}), 400

    row = db.execute("SELECT MAX(CAST(SUBSTR(id,4) AS INTEGER)) FROM employees").fetchone()
    next_id = (row[0] or 0) + 1
    emp_id = f"EMP{next_id:03d}"

    db.execute(
        "INSERT INTO employees(id, name, team, position, supervisor, entry_date, dongfu_id, work_hour_system, status) VALUES(?,?,?,?,?,?,?,?,?)",
        (emp_id, name, team, position, supervisor, entry_date, dongfu_id, work_hour_system, status)
    )
    db.commit()
    row = db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/employees/<emp_id>", methods=["PUT"])
def api_employees_update(emp_id):
    data = request.json
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        return jsonify({"error": "员工不存在"}), 404

    name = (data.get("name") or "").strip() or emp["name"]
    team = (data.get("team") or "").strip() or emp["team"]
    position = (data.get("position") or "").strip() or emp["position"]
    supervisor = data.get("supervisor") if "supervisor" in data else emp["supervisor"]
    if supervisor: supervisor = supervisor.strip()
    entry_date = data.get("entryDate") if "entryDate" in data else emp["entry_date"]
    if entry_date: entry_date = entry_date.strip()
    dongfu_id = data.get("dongfuId") if "dongfuId" in data else emp["dongfu_id"]
    if dongfu_id: dongfu_id = dongfu_id.strip()
    work_hour_system = data.get("workHourSystem") if "workHourSystem" in data else (emp["work_hour_system"] or "")
    if work_hour_system: work_hour_system = work_hour_system.strip()
    else: work_hour_system = ""
    status = (data.get("status") or "").strip() or emp["status"] or "active"

    if not name:
        return jsonify({"error": "请输入员工姓名"}), 400

    exist = db.execute("SELECT id FROM employees WHERE name=? AND id!=?", (name, emp_id)).fetchone()
    if exist:
        return jsonify({"error": f"员工\"{name}\"已存在"}), 400

    old_name = emp["name"]
    db.execute(
        "UPDATE employees SET name=?, team=?, position=?, supervisor=?, entry_date=?, dongfu_id=?, work_hour_system=?, status=?, updated_at=datetime('now','localtime') WHERE id=?",
        (name, team, position, supervisor or "", entry_date or "", dongfu_id or "", work_hour_system, status, emp_id)
    )
    # 如果改了名字，同步更新上下级引用和排班记录
    if old_name != name:
        db.execute("UPDATE employees SET supervisor=? WHERE supervisor=?", (name, old_name))
        db.execute("UPDATE schedules SET employee_name=? WHERE employee_name=?", (name, old_name))
    db.commit()
    row = db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/employees/import", methods=["POST"])
def api_employees_import():
    """Excel 批量导入员工"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传文件"}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file, read_only=True)
    except Exception:
        return jsonify({"error": "无法解析该文件，请上传 .xlsx 格式的 Excel（不支持旧版 .xls 格式）"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "Excel至少包含两行数据（表头+数据）"}), 400

    header = [str(c or "").strip() for c in rows[0]]
    name_idx = next((i for i, h in enumerate(header) if "姓名" in h), -1)
    team_idx = next((i for i, h in enumerate(header) if "团队" in h), -1)
    pos_idx = next((i for i, h in enumerate(header) if "岗位" in h), -1)
    sup_idx = next((i for i, h in enumerate(header) if "上级" in h), -1)
    date_idx = next((i for i, h in enumerate(header) if "入职" in h or "日期" in h), -1)
    dongfu_idx = next((i for i, h in enumerate(header) if "东福" in h or "工号" in h), -1)
    whs_idx = next((i for i, h in enumerate(header) if "工时" in h or "工时制度" in h), -1)
    status_idx = next((i for i, h in enumerate(header) if "状态" in h), -1)

    db = get_db()
    success = 0
    errors = []
    for rn, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        try:
            name = str(row[name_idx] or "").strip()
            team = str(row[team_idx] or "").strip()
            position = str(row[pos_idx] or "").strip()
        except IndexError:
            continue
        if not name or not team or not position:
            if name:
                errors.append(f"第{rn}行: 信息不完整")
            continue

        supervisor = ""
        entry_date = ""
        dongfu_id = ""
        work_hour_system = ""
        status = "active"
        if sup_idx >= 0:
            try:
                supervisor = str(row[sup_idx] or "").strip()
            except IndexError:
                pass
        if date_idx >= 0:
            try:
                raw_date = str(row[date_idx] or "").strip()
                entry_date = _parse_date(raw_date) or raw_date
            except IndexError:
                pass
        if dongfu_idx >= 0:
            try:
                dongfu_id = str(row[dongfu_idx] or "").strip()
            except IndexError:
                pass
        if whs_idx >= 0:
            try:
                work_hour_system = str(row[whs_idx] or "").strip()
            except IndexError:
                pass
        if status_idx >= 0:
            try:
                raw_status = str(row[status_idx] or "").strip()
                if raw_status in ("失效", "inactive", "0", "false", "否"):
                    status = "inactive"
            except IndexError:
                pass

        exist = db.execute("SELECT id FROM employees WHERE name=?", (name,)).fetchone()
        if exist:
            errors.append(f"第{rn}行: 员工\"{name}\"已存在，跳过")
            continue

        num_row = db.execute("SELECT MAX(CAST(SUBSTR(id,4) AS INTEGER)) FROM employees").fetchone()
        next_id = (num_row[0] or 0) + 1
        emp_id = f"EMP{next_id:03d}"

        db.execute(
            "INSERT INTO employees(id, name, team, position, supervisor, entry_date, dongfu_id, work_hour_system, status) VALUES(?,?,?,?,?,?,?,?,?)",
            (emp_id, name, team, position, supervisor, entry_date, dongfu_id, work_hour_system, status)
        )
        success += 1

    db.commit()
    wb.close()
    return jsonify({"ok": True, "success": success, "errors": errors})


@app.route("/api/employees/template", methods=["GET"])
def api_employees_template():
    """下载员工导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工导入模板"
    ws.append(["姓名", "东福工号", "团队", "岗位", "工时制度", "上级", "入职日期", "状态"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 10

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="员工导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/employees/<emp_id>", methods=["DELETE"])
def api_employees_delete(emp_id):
    db = get_db()
    emp = db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        return jsonify({"error": "员工不存在"}), 404

    subordinates = db.execute("SELECT name FROM employees WHERE supervisor=?", (emp["name"],)).fetchall()
    if subordinates:
        names = "、".join(r["name"] for r in subordinates)
        return jsonify({"error": f"无法删除：{names} 的上级为 {emp['name']}，请先调整"}), 400

    db.execute("DELETE FROM employees WHERE id=?", (emp_id,))
    db.commit()
    return jsonify({"ok": True})


# ==================== 班次 API ====================

@app.route("/api/shifts", methods=["GET"])
def api_shifts_list():
    db = get_db()
    rows = db.execute("SELECT * FROM shifts ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/shifts", methods=["POST"])
def api_shifts_create():
    data = request.json
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "请输入班次名称"}), 400

    db = get_db()
    exist = db.execute("SELECT id FROM shifts WHERE name=?", (name,)).fetchone()
    if exist:
        return jsonify({"error": f"班次\"{name}\"已存在"}), 400

    row = db.execute("SELECT MAX(CAST(SUBSTR(id,4) AS INTEGER)) FROM shifts").fetchone()
    next_id = (row[0] or 0) + 1
    shift_id = f"SHF{next_id:03d}"

    info = (data.get("info") or "").strip()
    start_time = (data.get("startTime") or "").strip()
    end_time = (data.get("endTime") or "").strip()
    lunch_start = (data.get("lunchStart") or "").strip()
    lunch_end = (data.get("lunchEnd") or "").strip()
    dinner_start = (data.get("dinnerStart") or "").strip()
    dinner_end = (data.get("dinnerEnd") or "").strip()
    work_hours = _calc_work_hours(start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end)

    db.execute(
        "INSERT INTO shifts(id, name, info, start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end, work_hours) VALUES(?,?,?,?,?,?,?,?,?,?)",
        (shift_id, name, info, start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end, work_hours)
    )
    db.commit()
    row = db.execute("SELECT * FROM shifts WHERE id=?", (shift_id,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/shifts/<shift_id>", methods=["PUT"])
def api_shifts_update(shift_id):
    data = request.json
    db = get_db()
    shift = db.execute("SELECT * FROM shifts WHERE id=?", (shift_id,)).fetchone()
    if not shift:
        return jsonify({"error": "班次不存在"}), 404

    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "请输入班次名称"}), 400

    exist = db.execute("SELECT id FROM shifts WHERE name=? AND id!=?", (name, shift_id)).fetchone()
    if exist:
        return jsonify({"error": f"班次\"{name}\"已存在"}), 400

    old_name = shift["name"]
    info = (data.get("info") or "").strip()
    start_time = (data.get("startTime") or "").strip()
    end_time = (data.get("endTime") or "").strip()
    lunch_start = (data.get("lunchStart") or "").strip()
    lunch_end = (data.get("lunchEnd") or "").strip()
    dinner_start = (data.get("dinnerStart") or "").strip()
    dinner_end = (data.get("dinnerEnd") or "").strip()
    work_hours = _calc_work_hours(start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end)

    db.execute(
        "UPDATE shifts SET name=?, info=?, start_time=?, end_time=?, lunch_start=?, lunch_end=?, dinner_start=?, dinner_end=?, work_hours=?, updated_at=datetime('now','localtime') WHERE id=?",
        (name, info, start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end, work_hours, shift_id)
    )
    # 如果改了名字，同步排班记录
    if old_name != name:
        db.execute("UPDATE schedules SET shift_name=? WHERE shift_name=?", (name, old_name))
    db.commit()
    row = db.execute("SELECT * FROM shifts WHERE id=?", (shift_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/shifts/<shift_id>", methods=["DELETE"])
def api_shifts_delete(shift_id):
    db = get_db()
    shift = db.execute("SELECT * FROM shifts WHERE id=?", (shift_id,)).fetchone()
    if not shift:
        return jsonify({"error": "班次不存在"}), 404
    db.execute("DELETE FROM shifts WHERE id=?", (shift_id,))
    db.commit()
    return jsonify({"ok": True})


# ==================== 排班 API ====================

@app.route("/api/schedules", methods=["GET"])
def api_schedules_list():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    db = get_db()
    if start and end:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM schedules WHERE schedule_date>=? AND schedule_date<=? ORDER BY schedule_date, employee_name",
            (start, end)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM schedules ORDER BY schedule_date, employee_name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/schedule/lookup", methods=["GET"])
def api_schedule_lookup():
    """查询某员工在某日期的班次信息"""
    employee = request.args.get("employee", "").strip()
    date = request.args.get("date", "").strip()
    if not employee or not date:
        return jsonify({"error": "缺少employee或date参数"}), 400
    db = get_db()
    row = db.execute(
        "SELECT s.shift_name, sh.start_time, sh.end_time FROM schedules s "
        "LEFT JOIN shifts sh ON sh.name = s.shift_name "
        "WHERE s.schedule_date = ? AND s.employee_name = ?",
        (date, employee)
    ).fetchone()
    if not row:
        return jsonify({"shift_name": None, "start_time": None, "end_time": None})
    return jsonify(dict(row))


@app.route("/api/schedules/batch", methods=["POST"])
def api_schedules_batch():
    """批量新增/更新排班: {entries: [{date, employee, shift}, ...]}"""
    data = request.json
    entries = data.get("entries") or []
    if not entries:
        return jsonify({"error": "无排班数据"}), 400

    db = get_db()
    count = 0
    for e in entries:
        d = (e.get("date") or "").strip()
        emp = (e.get("employee") or "").strip()
        shift = (e.get("shift") or "").strip()
        if not d or not emp:
            continue
        if not shift:
            shift = "休息"
        db.execute(
            "INSERT INTO schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name, created_at=datetime('now','localtime')",
            (d, emp, shift)
        )
        count += 1
    db.commit()
    return jsonify({"ok": True, "updated": count})


@app.route("/api/schedules/import", methods=["POST"])
def api_schedules_import():
    """Excel 导入排班"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传文件"}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"无法解析该文件：{e}"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "Excel至少包含两行数据"}), 400

    header = [str(c or "").lower() for c in rows[0]]
    date_idx = next((i for i, h in enumerate(header) if "日期" in h), -1)
    emp_idx = next((i for i, h in enumerate(header) if "员工" in h or "姓名" in h), -1)
    shift_idx = next((i for i, h in enumerate(header) if "班次" in h or "shift" in h), -1)

    if date_idx < 0 or emp_idx < 0 or shift_idx < 0:
        return jsonify({"error": "Excel需包含【日期,员工,班次】三列"}), 400

    db = get_db()
    count = 0
    failures = []
    for rn, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        try:
            d = str(row[date_idx] or "").strip()
            emp = str(row[emp_idx] or "").strip()
            shift = str(row[shift_idx] or "").strip()
        except IndexError:
            continue
        if not d or not emp:
            if emp or d:
                failures.append({"row": rn, "employee": emp or "(空)", "date": d or "(空)", "reason": "信息不完整"})
            continue
        if not shift:
            shift = "休息"
        date_str = _parse_date(d)
        if not date_str:
            failures.append({"row": rn, "employee": emp, "date": d, "reason": "日期格式无法识别"})
            continue
        db.execute(
            "INSERT INTO schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
            (date_str, emp, shift)
        )
        count += 1
    db.commit()
    wb.close()
    return jsonify({"ok": True, "updated": count, "failures": failures})


@app.route("/api/schedules/import-matrix", methods=["POST"])
def api_schedules_import_matrix():
    """Excel 矩阵格式导入排班
    格式：第一列为员工姓名，第一行为日期（从第二列开始），交叉单元格为班次代码
    A=A班, B=B班, C=C班, D=D班, E=E班, 假=请假, 空=休息
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传文件"}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"无法解析该文件：{e}"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "Excel至少包含表头和数据行"}), 400

    # 解析表头：第一列是"姓名"，其余列是日期
    raw_header = rows[0]
    header = [str(c or "").strip() for c in raw_header]
    dates = []
    for i in range(1, len(raw_header)):
        # 优先用原始值解析（可能是 datetime 对象），字符串版本作为备选
        d = _parse_date(raw_header[i]) or _parse_date(header[i])
        if d:
            dates.append((i, d))

    if not dates:
        sample = "、".join(header[1:6])
        return jsonify({"error": f"未识别到日期列，表头前几列为：{sample}。请使用日期格式（如2026-02-01、2/1、1日）"}), 400

    # 加载班次表，构建名称→名称的映射
    db = get_db()
    db_shifts = db.execute("SELECT name FROM shifts").fetchall()
    db_shift_names = [s["name"] for s in db_shifts]

    # 代码→班次名称 精确映射（优先）
    code_to_shift = {
        "A": "A班", "a": "A班",
        "B": "B班", "b": "B班",
        "C": "C班", "c": "C班",
        "T": "T班", "t": "T班",
        "E": "E班", "e": "E班",
        "F": "F班", "f": "F班",
        "休": "休息", "休假": "休息", "休息": "休息", "调休": "放休", "放休": "放休",
        "假": "请假", "请假": "请假",
    }

    def resolve_shift(raw_val):
        """将单元格值解析为班次名称。返回 (shift_name, None) 成功，或 (None, reason) 失败"""
        val = raw_val.strip()
        if not val:
            return "休息", None

        # 1. 精确映射
        if val in code_to_shift:
            target = code_to_shift[val]
            if target in db_shift_names:
                return target, None
            else:
                return None, f"映射目标班次\"{target}\"在班次表中不存在"

        # 2. 直接匹配班次名称
        if val in db_shift_names:
            return val, None

        # 3. 模糊匹配（双向子串）
        for sn in db_shift_names:
            if val in sn or sn in val:
                return sn, None

        return None, f"班次\"{val}\"无法识别"

    count = 0
    failures = []
    for rn, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        emp_name = str(row[0] or "").strip()
        if not emp_name:
            continue

        emp = db.execute("SELECT name FROM employees WHERE name=?", (emp_name,)).fetchone()
        if not emp:
            failures.append({"row": rn, "employee": emp_name, "date": "", "reason": "员工不存在"})
            continue

        for col_idx, date_str in dates:
            try:
                raw = str(row[col_idx] or "")
            except IndexError:
                raw = ""
            if not raw.strip():
                raw = ""  # 空值统一处理为休息

            shift, err = resolve_shift(raw)
            if err:
                failures.append({"row": rn, "employee": emp_name, "date": date_str, "reason": err})
                continue

            db.execute(
                "INSERT INTO schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
                (date_str, emp_name, shift)
            )
            count += 1

    db.commit()
    wb.close()
    return jsonify({"ok": True, "updated": count, "failures": failures})


@app.route("/api/schedules/export", methods=["GET"])
def api_schedules_export():
    """Excel 导出排班"""
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    db = get_db()

    # 查询员工
    employees = db.execute("SELECT name, team FROM employees ORDER BY id").fetchall()

    # 查询排班
    if start and end:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM schedules WHERE schedule_date>=? AND schedule_date<=? ORDER BY schedule_date",
            (start, end)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM schedules ORDER BY schedule_date"
        ).fetchall()

    # 构建 date -> emp -> shift 映射
    schedule_map = {}
    date_set = set()
    for r in rows:
        schedule_map[(r["schedule_date"], r["employee_name"])] = r["shift_name"]
        date_set.add(r["schedule_date"])

    wb = Workbook()
    ws = wb.active
    ws.title = "排班表"

    # 表头
    ws.cell(row=1, column=1, value="日期")
    ws.cell(row=1, column=2, value="员工")
    ws.cell(row=1, column=3, value="所属团队")
    ws.cell(row=1, column=4, value="班次")

    sorted_dates = sorted(date_set)
    row_idx = 2
    for d in sorted_dates:
        for emp in employees:
            shift = schedule_map.get((d, emp["name"]), "未排班")
            ws.cell(row=row_idx, column=1, value=d)
            ws.cell(row=row_idx, column=2, value=emp["name"])
            ws.cell(row=row_idx, column=3, value=emp["team"])
            ws.cell(row=row_idx, column=4, value=shift)
            row_idx += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="排班数据.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 月度排班 API ====================

@app.route("/api/monthly-schedules")
def api_monthly_schedules_list():
    """查询指定月份的所有月度排班记录"""
    year_month = request.args.get("year_month", "")
    try:
        y, m = year_month.split("-")
        y, m = int(y), int(m)
    except ValueError:
        return jsonify({"ok": False, "error": "参数 year_month 格式错误，需为 YYYY-MM"}), 400
    start_date = f"{y}-{m:02d}-01"
    import calendar as cal
    end_date = f"{y}-{m:02d}-{cal.monthrange(y, m)[1]}"

    db = get_db()
    rows = db.execute(
        "SELECT schedule_date, employee_name, shift_name, finalized FROM monthly_schedules WHERE schedule_date>=? AND schedule_date<=? ORDER BY schedule_date, employee_name",
        (start_date, end_date)
    ).fetchall()
    return jsonify({
        "ok": True,
        "data": [{"schedule_date": r["schedule_date"], "employee_name": r["employee_name"], "shift_name": r["shift_name"], "finalized": r["finalized"]} for r in rows],
        "finalized": any(r["finalized"] for r in rows)
    })


@app.route("/api/monthly-schedules", methods=["POST"])
def api_monthly_schedules_save():
    """保存单个单元格的班次（shift_name 为空则删除）"""
    data = request.get_json(silent=True) or {}
    schedule_date = (data.get("schedule_date") or "").strip()
    employee_name = (data.get("employee_name") or "").strip()
    shift_name = (data.get("shift_name") or "").strip()
    if not schedule_date or not employee_name:
        return jsonify({"ok": False, "error": "缺少必填字段"}), 400

    db = get_db()
    # 检查是否已 finalize
    y, m = schedule_date.split("-")[:2]
    locked = db.execute(
        "SELECT COUNT(*) as c FROM monthly_schedules WHERE schedule_date LIKE ? || '%' AND finalized=1",
        (f"{y}-{m}",)
    ).fetchone()["c"]
    if locked:
        return jsonify({"ok": False, "error": "该月已最终保存，不可修改"}), 403

    if not shift_name:
        db.execute("DELETE FROM monthly_schedules WHERE schedule_date=? AND employee_name=?",
                   (schedule_date, employee_name))
    else:
        if shift_name not in ("A班", "B班", "C班"):
            return jsonify({"ok": False, "error": f"不支持的班次: {shift_name}"}), 400
        db.execute(
            "INSERT INTO monthly_schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name, updated_at=datetime('now','localtime')",
            (schedule_date, employee_name, shift_name)
        )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/monthly-schedules/finalize", methods=["POST"])
def api_monthly_schedules_finalize():
    """最终保存：将当月数据导入 raw_schedules 并标记为已 finalize"""
    data = request.get_json(silent=True) or {}
    year_month = (data.get("year_month") or "").strip()
    try:
        y, m = year_month.split("-")
        y, m = int(y), int(m)
    except ValueError:
        return jsonify({"ok": False, "error": "参数 year_month 格式错误"}), 400

    db = get_db()
    # 检查是否已 finalize
    if db.execute("SELECT COUNT(*) as c FROM monthly_schedules WHERE schedule_date LIKE ? AND finalized=1",
                  (f"{y}-{m:02d}-%",)).fetchone()["c"] > 0:
        return jsonify({"ok": False, "error": "该月已最终保存，不可重复操作"}), 403

    start_date = f"{y}-{m:02d}-01"
    import calendar as cal
    end_date = f"{y}-{m:02d}-{cal.monthrange(y, m)[1]}"

    # 读取当月所有 monthly_schedules 记录
    rows = db.execute(
        "SELECT schedule_date, employee_name, shift_name FROM monthly_schedules WHERE schedule_date>=? AND schedule_date<=?",
        (start_date, end_date)
    ).fetchall()

    if not rows:
        return jsonify({"ok": False, "error": "该月无排班数据"}), 400

    # 删除当月已有的 raw_schedules
    db.execute("DELETE FROM raw_schedules WHERE schedule_date>=? AND schedule_date<=?",
               (start_date, end_date))
    # 同步删除对应的 schedules
    db.execute("DELETE FROM schedules WHERE schedule_date>=? AND schedule_date<=?",
               (start_date, end_date))

    count = 0
    for r in rows:
        db.execute(
            "INSERT INTO raw_schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
            (r["schedule_date"], r["employee_name"], r["shift_name"])
        )
        # 同步到 schedules
        db.execute(
            "INSERT INTO schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
            (r["schedule_date"], r["employee_name"], r["shift_name"])
        )
        count += 1

    # 标记为 finalized
    db.execute("UPDATE monthly_schedules SET finalized=1 WHERE schedule_date>=? AND schedule_date<=?",
               (start_date, end_date))
    db.commit()
    return jsonify({"ok": True, "count": count})


@app.route("/api/monthly-schedules/export")
def api_monthly_schedules_export():
    """导出月度排班 Excel，表格样式与页面一致"""
    year_month = request.args.get("year_month", "")
    try:
        y, m = year_month.split("-")
        y, m = int(y), int(m)
    except ValueError:
        return jsonify({"ok": False, "error": "参数 year_month 格式错误"}), 400

    import calendar as cal
    days = cal.monthrange(y, m)[1]
    start_date = f"{y}-{m:02d}-01"
    end_date = f"{y}-{m:02d}-{cal.monthrange(y, m)[1]}"

    db = get_db()

    # 读取员工（仅在线组/热线组/售后组，在职）
    all_emps = db.execute(
        "SELECT name, team FROM employees WHERE status='active' AND team IN ('在线组','热线组','售后组') ORDER BY team, name"
    ).fetchall()

    # 读取月度排班数据
    rows = db.execute(
        "SELECT schedule_date, employee_name, shift_name FROM monthly_schedules WHERE schedule_date>=? AND schedule_date<=?",
        (start_date, end_date)
    ).fetchall()
    sched_map = {}
    for r in rows:
        sched_map[(r["schedule_date"], r["employee_name"])] = r["shift_name"]

    # 读取班次工时
    shift_rows = db.execute("SELECT name, work_hours FROM shifts").fetchall()
    shift_hours = {sr["name"]: sr["work_hours"] for sr in shift_rows}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月排班"

    # 表头行1：日期
    headers = ["团队", "姓名"]
    for d in range(1, days + 1):
        headers.append(f"{m}/{d}")
    headers += ["排班工时", "排班天数", "A班天数", "B班天数", "C班天数"]
    ws.append(headers)

    # 表头行2：星期
    week_labels = ["日", "一", "二", "三", "四", "五", "六"]
    week_row = ["", ""]
    for d in range(1, days + 1):
        w = cal.weekday(y, m, d)  # 0=Mon...6=Sun
        week_row.append(week_labels[(w + 1) % 7])
    week_row += ["", "", "", "", ""]
    ws.append(week_row)

    # 数据行
    for emp in all_emps:
        row_data = [emp["team"], emp["name"]]
        a_count = b_count = c_count = 0
        total_hours = 0.0
        for d in range(1, days + 1):
            date_str = f"{y}-{m:02d}-{d:02d}"
            shift = sched_map.get((date_str, emp["name"]), "")
            row_data.append(shift)
            if shift == "A班":
                a_count += 1
                total_hours += shift_hours.get("A班", 8)
            elif shift == "B班":
                b_count += 1
                total_hours += shift_hours.get("B班", 11)
            elif shift == "C班":
                c_count += 1
                total_hours += shift_hours.get("C班", 8)
        total_days = a_count + b_count + c_count
        row_data += [f"{total_hours:.1f}h", f"{total_days}天", f"{a_count}天", f"{b_count}天", f"{c_count}天"]
        ws.append(row_data)

    # 汇总行
    summary_labels = ["当日上班人次", "当日A班人数", "当日B班人数", "当日C班人数"]
    for label in summary_labels:
        srow = [label, ""]
        for d in range(1, days + 1):
            date_str = f"{y}-{m:02d}-{d:02d}"
            count = 0
            for emp in all_emps:
                shift = sched_map.get((date_str, emp["name"]), "")
                if label == "当日上班人次":
                    if shift in ("A班", "B班", "C班"):
                        count += 1
                elif label == "当日A班人数" and shift == "A班":
                    count += 1
                elif label == "当日B班人数" and shift == "B班":
                    count += 1
                elif label == "当日C班人数" and shift == "C班":
                    count += 1
            srow.append(count if count > 0 else "")
        srow += ["", "", "", "", ""]
        ws.append(srow)

    # 设置列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    for d in range(1, days + 1):
        ws.column_dimensions[get_column_letter(d + 2)].width = 6

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"月度排班_{y}年{m}月.xlsx"
    )


# ==================== 原始排班 API ====================

def _sync_raw_to_schedule(db, date_str, emp_name, shift_name):
    """单向同步：原始排班 → 排班表"""
    db.execute(
        "INSERT INTO schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name, created_at=datetime('now','localtime')",
        (date_str, emp_name, shift_name)
    )


@app.route("/api/raw-schedules", methods=["GET"])
def api_raw_schedules_list():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    db = get_db()
    if start and end:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM raw_schedules WHERE schedule_date>=? AND schedule_date<=? ORDER BY schedule_date, employee_name",
            (start, end)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM raw_schedules ORDER BY schedule_date, employee_name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/raw-schedules/batch", methods=["POST"])
def api_raw_schedules_batch():
    """批量新增/更新原始排班: {entries: [{date, employee, shift}, ...]}，并单向同步到排班表"""
    data = request.json
    entries = data.get("entries") or []
    if not entries:
        return jsonify({"error": "无排班数据"}), 400

    db = get_db()
    count = 0
    for e in entries:
        d = (e.get("date") or "").strip()
        emp = (e.get("employee") or "").strip()
        shift = (e.get("shift") or "").strip()
        if not d or not emp:
            continue
        if not shift:
            shift = "休息"
        db.execute(
            "INSERT INTO raw_schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name, created_at=datetime('now','localtime')",
            (d, emp, shift)
        )
        _sync_raw_to_schedule(db, d, emp, shift)
        count += 1
    db.commit()
    return jsonify({"ok": True, "updated": count})


@app.route("/api/raw-schedules/import", methods=["POST"])
def api_raw_schedules_import():
    """Excel 导入原始排班（列表格式），单向同步到排班表"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传文件"}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"无法解析该文件：{e}"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "Excel至少包含两行数据"}), 400

    header = [str(c or "").lower() for c in rows[0]]
    date_idx = next((i for i, h in enumerate(header) if "日期" in h), -1)
    emp_idx = next((i for i, h in enumerate(header) if "员工" in h or "姓名" in h), -1)
    shift_idx = next((i for i, h in enumerate(header) if "班次" in h or "shift" in h), -1)

    if date_idx < 0 or emp_idx < 0 or shift_idx < 0:
        return jsonify({"error": "Excel需包含【日期,员工,班次】三列"}), 400

    db = get_db()
    count = 0
    failures = []
    for rn, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        try:
            d = str(row[date_idx] or "").strip()
            emp = str(row[emp_idx] or "").strip()
            shift = str(row[shift_idx] or "").strip()
        except IndexError:
            continue
        if not d or not emp:
            if emp or d:
                failures.append({"row": rn, "employee": emp or "(空)", "date": d or "(空)", "reason": "信息不完整"})
            continue
        if not shift:
            shift = "休息"
        date_str = _parse_date(d)
        if not date_str:
            failures.append({"row": rn, "employee": emp, "date": d, "reason": "日期格式无法识别"})
            continue
        db.execute(
            "INSERT INTO raw_schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
            (date_str, emp, shift)
        )
        _sync_raw_to_schedule(db, date_str, emp, shift)
        count += 1
    db.commit()
    wb.close()
    return jsonify({"ok": True, "updated": count, "failures": failures})


@app.route("/api/raw-schedules/import-matrix", methods=["POST"])
def api_raw_schedules_import_matrix():
    """Excel 矩阵格式导入原始排班，单向同步到排班表"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传文件"}), 400

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"无法解析该文件：{e}"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": "Excel至少包含表头和数据行"}), 400

    raw_header = rows[0]
    header = [str(c or "").strip() for c in raw_header]
    dates = []
    for i in range(1, len(raw_header)):
        d = _parse_date(raw_header[i]) or _parse_date(header[i])
        if d:
            dates.append((i, d))

    if not dates:
        sample = "、".join(header[1:6])
        return jsonify({"error": f"未识别到日期列，表头前几列为：{sample}。请使用日期格式（如2026-02-01、2/1、1日）"}), 400

    db = get_db()
    db_shifts = db.execute("SELECT name FROM shifts").fetchall()
    db_shift_names = [s["name"] for s in db_shifts]

    code_to_shift = {
        "A": "A班", "a": "A班",
        "B": "B班", "b": "B班",
        "C": "C班", "c": "C班",
        "T": "T班", "t": "T班",
        "E": "E班", "e": "E班",
        "F": "F班", "f": "F班",
        "休": "休息", "休假": "休息", "休息": "休息", "调休": "放休", "放休": "放休",
        "假": "请假", "请假": "请假",
    }

    def resolve_shift(raw_val):
        val = raw_val.strip()
        if not val:
            return "休息", None
        if val in code_to_shift:
            target = code_to_shift[val]
            if target in db_shift_names:
                return target, None
            else:
                return None, f"映射目标班次\"{target}\"在班次表中不存在"
        if val in db_shift_names:
            return val, None
        for sn in db_shift_names:
            if val in sn or sn in val:
                return sn, None
        return None, f"班次\"{val}\"无法识别"

    count = 0
    failures = []
    for rn, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        emp_name = str(row[0] or "").strip()
        if not emp_name:
            continue

        emp = db.execute("SELECT name FROM employees WHERE name=?", (emp_name,)).fetchone()
        if not emp:
            failures.append({"row": rn, "employee": emp_name, "date": "", "reason": "员工不存在"})
            continue

        for col_idx, date_str in dates:
            try:
                raw = str(row[col_idx] or "")
            except IndexError:
                raw = ""
            if not raw.strip():
                raw = ""

            shift, err = resolve_shift(raw)
            if err:
                failures.append({"row": rn, "employee": emp_name, "date": date_str, "reason": err})
                continue

            db.execute(
                "INSERT INTO raw_schedules(schedule_date, employee_name, shift_name) VALUES(?,?,?) ON CONFLICT(schedule_date, employee_name) DO UPDATE SET shift_name=excluded.shift_name",
                (date_str, emp_name, shift)
            )
            _sync_raw_to_schedule(db, date_str, emp_name, shift)
            count += 1

    db.commit()
    wb.close()
    return jsonify({"ok": True, "updated": count, "failures": failures})


@app.route("/api/raw-schedules/export", methods=["GET"])
def api_raw_schedules_export():
    """Excel 导出原始排班"""
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    db = get_db()

    employees = db.execute("SELECT name, team FROM employees ORDER BY id").fetchall()

    if start and end:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM raw_schedules WHERE schedule_date>=? AND schedule_date<=? ORDER BY schedule_date",
            (start, end)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT schedule_date, employee_name, shift_name FROM raw_schedules ORDER BY schedule_date"
        ).fetchall()

    schedule_map = {}
    date_set = set()
    for r in rows:
        schedule_map[(r["schedule_date"], r["employee_name"])] = r["shift_name"]
        date_set.add(r["schedule_date"])

    wb = Workbook()
    ws = wb.active
    ws.title = "原始排班表"

    ws.cell(row=1, column=1, value="日期")
    ws.cell(row=1, column=2, value="员工")
    ws.cell(row=1, column=3, value="所属团队")
    ws.cell(row=1, column=4, value="班次")

    sorted_dates = sorted(date_set)
    row_idx = 2
    for d in sorted_dates:
        for emp in employees:
            shift = schedule_map.get((d, emp["name"]), "未排班")
            ws.cell(row=row_idx, column=1, value=d)
            ws.cell(row=row_idx, column=2, value=emp["name"])
            ws.cell(row=row_idx, column=3, value=emp["team"])
            ws.cell(row=row_idx, column=4, value=shift)
            row_idx += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="原始排班数据.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 工作安排 API ====================

@app.route("/api/assignments", methods=["GET"])
def api_assignments_list():
    """查询某日工作安排"""
    date_str = request.args.get("date", "").strip()
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    db = get_db()
    rows = db.execute(
        "SELECT * FROM daily_assignments WHERE date=? ORDER BY employee_name",
        (date_str,)
    ).fetchall()
    results = []
    for r in rows:
        d = dict(r)
        d["work_types"] = json.loads(d["work_types"])
        results.append(d)
    return jsonify(results)


@app.route("/api/assignments", methods=["POST"])
def api_assignments_save():
    """批量保存工作安排: {date, assignments: [{employee_name, work_types, lunch_slot, dinner_slot}, ...]}"""
    data = request.json
    date_str = (data.get("date") or "").strip()
    entries = data.get("assignments") or []
    if not date_str:
        return jsonify({"error": "日期不能为空"}), 400

    db = get_db()
    count = 0
    for e in entries:
        emp = (e.get("employee_name") or "").strip()
        work_types = json.dumps(e.get("work_types") or [], ensure_ascii=False)
        lunch_slot = (e.get("lunch_slot") or "").strip()
        dinner_slot = (e.get("dinner_slot") or "").strip()
        if not emp:
            continue
        db.execute(
            "INSERT INTO daily_assignments(date, employee_name, work_types, lunch_slot, dinner_slot) "
            "VALUES(?,?,?,?,?) ON CONFLICT(date, employee_name) DO UPDATE SET "
            "work_types=excluded.work_types, lunch_slot=excluded.lunch_slot, dinner_slot=excluded.dinner_slot, "
            "updated_at=datetime('now','localtime')",
            (date_str, emp, work_types, lunch_slot, dinner_slot)
        )
        count += 1
    db.commit()
    return jsonify({"ok": True, "updated": count})


# ==================== 排班调整 API ====================

def _calc_leave_hours(start_time, end_time):
    """计算加班时长（小时），支持跨天"""
    sm = _time_to_minutes(start_time)
    em = _time_to_minutes(end_time)
    if sm is None or em is None:
        return 0
    if em <= sm:
        em += 24 * 60
    return round((em - sm) / 6) / 10


def _validate_rest_hours(db, employee_name, leave_date, new_hours, exclude_id=None):
    """校验放休总时长不超过当日排班班次的工作时长"""
    # 查询当日排班班次的工作时长
    row = db.execute(
        "SELECT s.shift_name, sh.work_hours FROM schedules s LEFT JOIN shifts sh ON s.shift_name=sh.name WHERE s.schedule_date=? AND s.employee_name=?",
        (leave_date, employee_name)
    ).fetchone()
    if not row or not row["work_hours"] or row["work_hours"] <= 0:
        return  # 无有效班次，不校验
    shift_hours = row["work_hours"]

    # 查询当日已有放休+换休总时长（排除当前编辑的记录）
    sql = "SELECT COALESCE(SUM(hours),0) FROM leave_records WHERE type IN ('rest','换休') AND leave_date=? AND employee_name=?"
    params = [leave_date, employee_name]
    if exclude_id is not None:
        sql += " AND id != ?"
        params.append(exclude_id)
    existing_rest = db.execute(sql, params).fetchone()[0]

    total_rest = existing_rest + new_hours
    if total_rest > shift_hours:
        raise ValueError(f"休息时长({total_rest:.1f}h)超出当日班次工作时长({shift_hours:.1f}h)")


@app.route("/api/leave-records", methods=["GET"])
def api_leave_records_list():
    team = request.args.get("team", "").strip()
    month = request.args.get("month", "").strip()
    record_type = request.args.get("type", "").strip()
    db = get_db()

    sql = "SELECT * FROM leave_records WHERE 1=1"
    params = []
    if team:
        sql += " AND team = ?"
        params.append(team)
    if month:
        sql += " AND leave_date LIKE ?"
        params.append(month + "%")
    if record_type:
        sql += " AND type = ?"
        params.append(record_type)
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/leave-records", methods=["POST"])
def api_leave_records_create():
    data = request.json
    record_type = (data.get("type") or "overtime").strip()
    if record_type not in ("overtime", "leave", "rest", "换班", "换休"):
        return jsonify({"error": "无效的记录类型"}), 400
    is_leave = record_type == "leave"

    team = (data.get("team") or "").strip()
    leave_date = (data.get("leaveDate") or "").strip()
    employee_name = (data.get("employeeName") or "").strip()
    start_time = (data.get("startTime") or "").strip()
    end_time = (data.get("endTime") or "").strip()
    remark = (data.get("remark") or "").strip()
    submitter = (data.get("submitter") or "").strip()
    deduction = data.get("deduction", 0)
    if deduction is None:
        deduction = 0
    deduction = float(deduction)

    if not team:
        return jsonify({"error": "请选择所属团队"}), 400
    if not leave_date:
        return jsonify({"error": "请选择日期"}), 400
    if not employee_name:
        return jsonify({"error": "请选择员工"}), 400
    if not is_leave and not start_time:
        return jsonify({"error": "请选择开始时间"}), 400
    if not is_leave and not end_time:
        return jsonify({"error": "请选择结束时间"}), 400
    if len(remark) > 200:
        return jsonify({"error": "备注不能超过200字符"}), 400
    if deduction < 0 or deduction > 8:
        return jsonify({"error": "扣除时长范围0~8小时"}), 400

    db = get_db()

    # 同人同日唯一校验
    existing = db.execute(
        "SELECT id FROM leave_records WHERE employee_name=? AND leave_date=? LIMIT 1",
        (employee_name, leave_date)
    ).fetchone()
    if existing:
        return jsonify({"error": "员工当日已有调班，请确认后再提交"}), 400

    emp = db.execute("SELECT dongfu_id FROM employees WHERE name=?", (employee_name,)).fetchone()
    dongfu_id = emp["dongfu_id"] if emp else ""

    if is_leave:
        hours = 0
    else:
        hours = max(0, round((_calc_leave_hours(start_time, end_time) - deduction) * 10) / 10)

    # 放休/换休时长校验：总休息时长不能超过班次工作时长
    if record_type in ("rest", "换休"):
        try:
            _validate_rest_hours(db, employee_name, leave_date, hours, exclude_id=None)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

    db.execute(
        "INSERT INTO leave_records(type, team, leave_date, dongfu_id, employee_name, start_time, end_time, hours, remark, submitter, deduction) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        (record_type, team, leave_date, dongfu_id, employee_name, start_time, end_time, hours, remark, submitter, deduction)
    )
    db.commit()
    row = db.execute("SELECT * FROM leave_records WHERE id=last_insert_rowid()").fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/leave-records/<int:record_id>", methods=["PUT"])
def api_leave_records_update(record_id):
    data = request.json
    db = get_db()
    record = db.execute("SELECT * FROM leave_records WHERE id=?", (record_id,)).fetchone()
    if not record:
        return jsonify({"error": "记录不存在"}), 404

    record_type = (data.get("type") or (record["type"] if record["type"] else None) or "overtime").strip()
    if record_type not in ("overtime", "leave", "rest", "换班", "换休"):
        return jsonify({"error": "无效的记录类型"}), 400
    is_leave = record_type == "leave"
    was_leave = (record["type"] or "overtime") == "leave"

    team = (data.get("team") or "").strip()
    leave_date = (data.get("leaveDate") or "").strip()
    employee_name = (data.get("employeeName") or "").strip()
    start_time = (data.get("startTime") or "").strip()
    end_time = (data.get("endTime") or "").strip()
    remark = (data.get("remark") or "").strip()
    submitter = (data.get("submitter") or "").strip()
    deduction = data.get("deduction", 0)
    if deduction is None:
        deduction = 0
    deduction = float(deduction)

    if not team or not leave_date or not employee_name:
        return jsonify({"error": "必填字段不能为空"}), 400
    if not is_leave and (not start_time or not end_time):
        return jsonify({"error": "必填字段不能为空"}), 400
    if len(remark) > 200:
        return jsonify({"error": "备注不能超过200字符"}), 400
    if deduction < 0 or deduction > 8:
        return jsonify({"error": "扣除时长范围0~8小时"}), 400

    # 同人同日唯一校验（排除当前记录自身）
    existing = db.execute(
        "SELECT id FROM leave_records WHERE employee_name=? AND leave_date=? AND id!=? LIMIT 1",
        (employee_name, leave_date, record_id)
    ).fetchone()
    if existing:
        return jsonify({"error": "员工当日已有调班，请确认后再提交"}), 400

    emp = db.execute("SELECT dongfu_id FROM employees WHERE name=?", (employee_name,)).fetchone()
    dongfu_id = emp["dongfu_id"] if emp else ""

    if is_leave:
        hours = 0
    else:
        hours = max(0, round((_calc_leave_hours(start_time, end_time) - deduction) * 10) / 10)

    # 放休/换休时长校验：总休息时长不能超过班次工作时长
    if record_type in ("rest", "换休"):
        try:
            _validate_rest_hours(db, employee_name, leave_date, hours, exclude_id=record_id)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

    db.execute(
        "UPDATE leave_records SET type=?, team=?, leave_date=?, dongfu_id=?, employee_name=?, start_time=?, end_time=?, hours=?, remark=?, submitter=?, deduction=?, updated_at=datetime('now','localtime') WHERE id=?",
        (record_type, team, leave_date, dongfu_id, employee_name, start_time, end_time, hours, remark, submitter, deduction, record_id)
    )
    db.commit()
    row = db.execute("SELECT * FROM leave_records WHERE id=?", (record_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/leave-records/<int:record_id>", methods=["DELETE"])
def api_leave_records_delete(record_id):
    db = get_db()
    record = db.execute("SELECT * FROM leave_records WHERE id=?", (record_id,)).fetchone()
    if not record:
        return jsonify({"error": "记录不存在"}), 404
    db.execute("DELETE FROM leave_records WHERE id=?", (record_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/leave-records/export", methods=["GET"])
def api_leave_records_export():
    team = request.args.get("team", "").strip()
    month = request.args.get("month", "").strip()
    employees_str = request.args.get("employees", "").strip()
    db = get_db()

    sql = "SELECT * FROM leave_records WHERE 1=1"
    params = []
    if team:
        sql += " AND team = ?"
        params.append(team)
    if month:
        sql += " AND leave_date LIKE ?"
        params.append(month + "%")
    if employees_str:
        names = [n.strip() for n in employees_str.split(",") if n.strip()]
        if names:
            placeholders = ",".join("?" for _ in names)
            sql += f" AND employee_name IN ({placeholders})"
            params.extend(names)
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "排班调整记录"
    type_map = {"overtime": "加班", "leave": "请假", "rest": "放休", "换班": "换班", "换休": "换休"}
    ws.append(["清单ID", "类型", "所属团队", "日期", "东福工号", "员工", "开始时间", "结束时间", "时长(h)", "备注", "提交人", "最后修改时间"])

    for r in rows:
        leave_date = r["leave_date"]
        start_time = r["start_time"] or ""
        end_time = r["end_time"] or ""
        r_type = r["type"] or "overtime"

        if r_type == "leave":
            export_start = ""
            export_end = ""
        else:
            export_start = f"{leave_date} {start_time}"
            if _time_to_minutes(end_time) is not None and _time_to_minutes(start_time) is not None:
                if _time_to_minutes(end_time) <= _time_to_minutes(start_time):
                    from datetime import datetime as dt, timedelta
                    next_day = dt.strptime(leave_date, "%Y-%m-%d") + timedelta(days=1)
                    export_end = f"{next_day.strftime('%Y-%m-%d')} {end_time}"
                else:
                    export_end = f"{leave_date} {end_time}"
            else:
                export_end = f"{leave_date} {end_time}"

        ws.append([
            r["id"], type_map.get(r_type, r_type), r["team"], r["leave_date"],
            r["dongfu_id"], r["employee_name"],
            export_start, export_end, r["hours"], r["remark"], r["submitter"],
            r["updated_at"] or r["created_at"]
        ])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 20

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="排班调整记录导出.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 换班/换休 API ====================

@app.route("/api/swap-records", methods=["GET"])
def api_swap_records_list():
    team = request.args.get("team", "").strip()
    month = request.args.get("month", "").strip()
    db = get_db()
    sql = "SELECT * FROM swap_records WHERE 1=1"
    params = []
    if team:
        sql += " AND (team_a = ? OR team_b = ?)"
        params.extend([team, team])
    if month:
        sql += " AND (date_a LIKE ? OR date_b LIKE ?)"
        params.extend([month + "%", month + "%"])
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/swap-records", methods=["POST"])
def api_swap_records_create():
    data = request.json
    swap_type = (data.get("swapType") or "").strip()
    person_a = (data.get("personA") or "").strip()
    date_a = (data.get("dateA") or "").strip()
    person_b = (data.get("personB") or "").strip()
    date_b = (data.get("dateB") or "").strip()
    remark = (data.get("remark") or "").strip()
    operator = (data.get("operator") or "").strip()

    if swap_type not in ("shift_swap", "rest_swap"):
        return jsonify({"error": "无效的换班类型"}), 400
    if not person_a or not date_a:
        return jsonify({"error": "请选择换班人和日期"}), 400
    if not person_b or not date_b:
        return jsonify({"error": "请选择交换人和日期"}), 400
    if not remark:
        return jsonify({"error": "请填写备注"}), 400
    if len(remark) > 200:
        return jsonify({"error": "备注不能超过200字符"}), 400
    if not operator:
        return jsonify({"error": "请选择操作人"}), 400

    db = get_db()

    # 查询两个日期上两人的排班
    row_a1 = db.execute(
        "SELECT shift_name FROM schedules WHERE schedule_date=? AND employee_name=?",
        (date_a, person_a)
    ).fetchone()
    row_a2 = db.execute(
        "SELECT shift_name FROM schedules WHERE schedule_date=? AND employee_name=?",
        (date_b, person_a)
    ).fetchone()
    row_b1 = db.execute(
        "SELECT shift_name FROM schedules WHERE schedule_date=? AND employee_name=?",
        (date_a, person_b)
    ).fetchone()
    row_b2 = db.execute(
        "SELECT shift_name FROM schedules WHERE schedule_date=? AND employee_name=?",
        (date_b, person_b)
    ).fetchone()

    if not row_a1:
        return jsonify({"error": f"{person_a} 在 {date_a} 无排班记录"}), 400
    if not row_b2:
        return jsonify({"error": f"{person_b} 在 {date_b} 无排班记录"}), 400

    # 查询团队
    emp_a = db.execute("SELECT team FROM employees WHERE name=?", (person_a,)).fetchone()
    team_a = emp_a["team"] if emp_a else ""

    if person_a == person_b:
        # 自换：同一人两个日期互换
        if not row_a2:
            return jsonify({"error": f"{person_a} 在 {date_b} 无排班记录"}), 400
        a1_old = row_a1["shift_name"]
        a2_old = row_a2["shift_name"]
        team_b = team_a

        cur = db.execute(
            "INSERT INTO swap_records(swap_type, person_a, team_a, date_a, shift_a, person_b, team_b, date_b, shift_b, remark, operator) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (swap_type, person_a, team_a, date_a, a1_old, person_a, team_a, date_b, a2_old, remark, operator)
        )
        swap_id = cur.lastrowid

        # 换班记录：date_b 替自己原 date_a 的班
        _create_swap_overtime(db, person_a, team_a, date_b, a1_old, person_a, operator, remark)
        # 换休记录：date_a 原班次不上了
        _create_swap_rest(db, person_a, team_a, date_a, a1_old, person_a, operator, remark)
        # 换班记录：date_a 替自己原 date_b 的班
        _create_swap_overtime(db, person_a, team_a, date_a, a2_old, person_a, operator, remark)
        # 换休记录：date_b 原班次不上了
        _create_swap_rest(db, person_a, team_a, date_b, a2_old, person_a, operator, remark)
    else:
        # 互换算两人在两个日期上的班次
        a1_old = row_a1["shift_name"]
        a2_old = row_a2["shift_name"] if row_a2 else "休息"
        b1_old = row_b1["shift_name"] if row_b1 else "休息"
        b2_old = row_b2["shift_name"]

        emp_b = db.execute("SELECT team FROM employees WHERE name=?", (person_b,)).fetchone()
        team_b = emp_b["team"] if emp_b else ""

        cur = db.execute(
            "INSERT INTO swap_records(swap_type, person_a, team_a, date_a, shift_a, person_b, team_b, date_b, shift_b, remark, operator) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (swap_type, person_a, team_a, date_a, a1_old, person_b, team_b, date_b, b2_old, remark, operator)
        )
        swap_id = cur.lastrowid

        # 换休：各人在自己有工作班次的日期上创建
        _create_swap_rest(db, person_a, team_a, date_a, a1_old, person_b, operator, remark)
        _create_swap_rest(db, person_a, team_a, date_b, a2_old, person_b, operator, remark)
        _create_swap_rest(db, person_b, team_b, date_a, b1_old, person_a, operator, remark)
        _create_swap_rest(db, person_b, team_b, date_b, b2_old, person_a, operator, remark)
        # 换班：各人在对方有工作班次的日期上替班
        _create_swap_overtime(db, person_a, team_a, date_a, b1_old, person_b, operator, remark)
        _create_swap_overtime(db, person_a, team_a, date_b, b2_old, person_b, operator, remark)
        _create_swap_overtime(db, person_b, team_b, date_a, a1_old, person_a, operator, remark)
        _create_swap_overtime(db, person_b, team_b, date_b, a2_old, person_a, operator, remark)

    db.commit()
    row = db.execute("SELECT * FROM swap_records WHERE id=?", (swap_id,)).fetchone()
    return jsonify(dict(row)), 201


def _create_swap_overtime(db, employee, team, date, shift_name, other_person, operator, remark=None):
    """创建换班记录（替别人上班），类型=换班"""
    if shift_name in ("休息", "放休", "请假", ""):
        return

    shift = db.execute(
        "SELECT start_time, end_time, work_hours FROM shifts WHERE name=?",
        (shift_name,)
    ).fetchone()
    if not shift or not shift["start_time"] or not shift["end_time"]:
        return

    start_time = shift["start_time"]
    end_time = shift["end_time"]
    hours = shift["work_hours"] if shift["work_hours"] else _calc_leave_hours(start_time, end_time)

    emp = db.execute("SELECT dongfu_id FROM employees WHERE name=?", (employee,)).fetchone()
    dongfu_id = emp["dongfu_id"] if emp else ""

    swap_remark = remark if remark else f"与{other_person}换班"

    db.execute(
        "INSERT INTO leave_records(type, team, leave_date, dongfu_id, employee_name, start_time, end_time, hours, remark, submitter) "
        "VALUES('换班',?,?,?,?,?,?,?,?,?)",
        (team, date, dongfu_id, employee, start_time, end_time, hours, swap_remark, operator)
    )


def _create_swap_rest(db, employee, team, date, shift_name, other_person, operator, remark=None):
    """创建换休记录（原班不上，效果=放休），类型=换休"""
    if shift_name in ("休息", "放休", "请假", ""):
        return

    shift = db.execute(
        "SELECT start_time, end_time, work_hours FROM shifts WHERE name=?",
        (shift_name,)
    ).fetchone()
    if not shift or not shift["start_time"] or not shift["end_time"]:
        return

    start_time = shift["start_time"]
    end_time = shift["end_time"]
    hours = shift["work_hours"] if shift["work_hours"] else _calc_leave_hours(start_time, end_time)

    emp = db.execute("SELECT dongfu_id FROM employees WHERE name=?", (employee,)).fetchone()
    dongfu_id = emp["dongfu_id"] if emp else ""

    rest_remark = remark if remark else f"与{other_person}换班"

    db.execute(
        "INSERT INTO leave_records(type, team, leave_date, dongfu_id, employee_name, start_time, end_time, hours, remark, submitter) "
        "VALUES('换休',?,?,?,?,?,?,?,?,?)",
        (team, date, dongfu_id, employee, start_time, end_time, hours, rest_remark, operator)
    )


# ==================== 辅助函数 ====================

def _time_to_minutes(t):
    if not t:
        return None
    parts = t.split(":")
    if len(parts) != 2:
        return None
    return int(parts[0]) * 60 + int(parts[1])


def _calc_work_hours(start_time, end_time, lunch_start, lunch_end, dinner_start, dinner_end):
    sm = _time_to_minutes(start_time)
    em = _time_to_minutes(end_time)
    if sm is None or em is None:
        return 0
    if em <= sm:
        em += 24 * 60
    work_min = em - sm
    if lunch_start and lunch_end:
        ls = _time_to_minutes(lunch_start)
        le = _time_to_minutes(lunch_end)
        if ls is not None and le is not None:
            if le <= ls:
                le += 24 * 60
            work_min -= (le - ls)
    if dinner_start and dinner_end:
        ds = _time_to_minutes(dinner_start)
        de = _time_to_minutes(dinner_end)
        if ds is not None and de is not None:
            if de <= ds:
                de += 24 * 60
            work_min -= (de - ds)
    return round(work_min / 6) / 10


def _parse_date(val):
    """尝试解析各种日期格式为 YYYY-MM-DD"""
    import re
    # 处理 datetime 对象（openpyxl 可能返回）
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    # YYYY-MM-DD（可能带时间部分，如 "2026-05-01 00:00:00"）
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})(?:\s|T|$)", val)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    # YYYY/MM/DD
    if re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", val):
        return val.replace("/", "-")
    # YYYY.M.DD 或 YYYY.M.D
    m = re.match(r"^(\d{4})\.(\d{1,2})\.(\d{1,2})$", val)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    # YYYY年M月D日
    m = re.match(r"^(\d{4})年(\d{1,2})月(\d{1,2})日$", val)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    # MM/DD/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", val)
    if m:
        return f"{m.group(3)}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
    # M/D（月/日，如 2/1、5/01）
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", val)
    if m:
        now = datetime.now()
        return f"{now.year}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
    # M月D日（如 5月1日、12月15日）
    m = re.match(r"^(\d{1,2})月(\d{1,2})日$", val)
    if m:
        now = datetime.now()
        return f"{now.year}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
    # D日（如 1日、15日）
    m = re.match(r"^(\d{1,2})日$", val)
    if m:
        now = datetime.now()
        return f"{now.year}-{now.month:02d}-{m.group(1).zfill(2)}"
    # 纯数字 1-31（日数）
    try:
        n = int(val)
        if 1 <= n <= 31:
            now = datetime.now()
            return f"{now.year}-{now.month:02d}-{n:02d}"
    except (ValueError, TypeError):
        pass
    # Excel 序列号
    try:
        n = float(val)
        if n > 40000:
            from datetime import timedelta
            d = datetime(1899, 12, 30) + timedelta(days=n)
            return d.strftime("%Y-%m-%d")
    except ValueError:
        pass
    return ""


# ==================== 工作安排统计 API ====================

ASSIGNMENT_WORK_TYPES = ["反向工单", "自主售后", "售后单", "紧急", "本地生活"]


@app.route("/api/assignment-stats", methods=["GET"])
def api_assignment_stats():
    """按自然月统计工作安排，按团队过滤"""
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    team_str = request.args.get("team", "售后组").strip()

    now = datetime.now()
    if not year:
        year = str(now.year)
    if not month:
        month = str(now.month)

    try:
        year_int = int(year)
        month_int = int(month)
    except (ValueError, TypeError):
        return jsonify({"error": "年份或月份格式无效"}), 400
    last_day = calendar.monthrange(year_int, month_int)[1]
    start_date = f"{year_int}-{month_int:02d}-01"
    end_date = f"{year_int}-{month_int:02d}-{last_day}"

    db = get_db()

    # 支持逗号分隔多团队
    teams = [t.strip() for t in team_str.split(",") if t.strip()]

    # 获取指定团队的员工
    if teams:
        placeholders = ",".join("?" for _ in teams)
        employees = db.execute(
            f"SELECT name FROM employees WHERE team IN ({placeholders}) AND status='active' ORDER BY name",
            teams
        ).fetchall()
    else:
        employees = db.execute(
            "SELECT name FROM employees WHERE status='active' ORDER BY name"
        ).fetchall()
    emp_names = {e["name"] for e in employees}

    # 查询当月所有工作安排
    rows = db.execute(
        "SELECT employee_name, work_types FROM daily_assignments WHERE date>=? AND date<=? ORDER BY date",
        (start_date, end_date)
    ).fetchall()

    # 统计：员工 -> 工作类型 -> 次数
    stats = {}
    for e in employees:
        stats[e["name"]] = {w: 0 for w in ASSIGNMENT_WORK_TYPES}

    for r in rows:
        emp = r["employee_name"]
        if emp not in emp_names:
            continue
        try:
            work_types = json.loads(r["work_types"])
        except (json.JSONDecodeError, TypeError):
            continue
        for w in work_types:
            if w in stats[emp]:
                stats[emp][w] += 1

    results = []
    for e in employees:
        entry = {"name": e["name"]}
        entry.update(stats[e["name"]])
        results.append(entry)

    # 所有可选团队
    all_teams = db.execute("SELECT DISTINCT team FROM employees WHERE status='active' ORDER BY team").fetchall()
    team_list = [r["team"] for r in all_teams]

    return jsonify({"rows": results, "teams": team_list, "month": f"{year_int}年{month_int}月"})


@app.route("/api/assignment-stats/export", methods=["GET"])
def api_assignment_stats_export():
    """导出工作安排统计Excel"""
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    team_str = request.args.get("team", "售后组").strip()

    now = datetime.now()
    if not year:
        year = str(now.year)
    if not month:
        month = str(now.month)

    try:
        year_int = int(year)
        month_int = int(month)
    except (ValueError, TypeError):
        return jsonify({"error": "年份或月份格式无效"}), 400
    last_day = calendar.monthrange(year_int, month_int)[1]
    start_date = f"{year_int}-{month_int:02d}-01"
    end_date = f"{year_int}-{month_int:02d}-{last_day}"

    db = get_db()

    teams = [t.strip() for t in team_str.split(",") if t.strip()]
    if teams:
        placeholders = ",".join("?" for _ in teams)
        employees = db.execute(
            f"SELECT name FROM employees WHERE team IN ({placeholders}) AND status='active' ORDER BY name",
            teams
        ).fetchall()
    else:
        employees = db.execute(
            "SELECT name FROM employees WHERE status='active' ORDER BY name"
        ).fetchall()
    emp_names = {e["name"] for e in employees}

    rows = db.execute(
        "SELECT employee_name, work_types FROM daily_assignments WHERE date>=? AND date<=? ORDER BY date",
        (start_date, end_date)
    ).fetchall()

    stats = {}
    for e in employees:
        stats[e["name"]] = {w: 0 for w in ASSIGNMENT_WORK_TYPES}

    for r in rows:
        emp = r["employee_name"]
        if emp not in emp_names:
            continue
        try:
            work_types = json.loads(r["work_types"])
        except (json.JSONDecodeError, TypeError):
            continue
        for w in work_types:
            if w in stats[emp]:
                stats[emp][w] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year_int}年{month_int}月工作安排统计"
    ws.append(["员工姓名"] + ASSIGNMENT_WORK_TYPES)

    for e in employees:
        row = [e["name"]] + [stats[e["name"]][w] for w in ASSIGNMENT_WORK_TYPES]
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    for i in range(len(ASSIGNMENT_WORK_TYPES)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name=f"工作安排统计_{year_int}年{month_int}月.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 月度时长统计 API ====================

@app.route("/api/monthly-hours-stats", methods=["GET"])
def api_monthly_hours_stats():
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    work_system = request.args.get("work_system", "").strip()
    teams_str = request.args.get("teams", "").strip()

    now = datetime.now()
    if not year:
        year = str(now.year)
    if not month:
        month = str(now.month)

    try:
        year_int = int(year)
        month_int = int(month)
    except (ValueError, TypeError):
        return jsonify({"error": "年份或月份格式无效"}), 400
    last_day = calendar.monthrange(year_int, month_int)[1]
    start_date = f"{year_int}-{month_int:02d}-01"
    end_date = f"{year_int}-{month_int:02d}-{last_day}"

    db = get_db()

    query = "SELECT id, name, team, dongfu_id, work_hour_system FROM employees WHERE status='active'"
    params = []
    if work_system:
        query += " AND work_hour_system = ?"
        params.append(work_system)
    teams = [t.strip() for t in teams_str.split(",") if t.strip()]
    if teams:
        placeholders = ",".join("?" for _ in teams)
        query += f" AND team IN ({placeholders})"
        params.extend(teams)
    query += " ORDER BY id"
    employees = db.execute(query, params).fetchall()

    shifts = db.execute("SELECT id, name, work_hours FROM shifts").fetchall()
    shift_map = {s["name"]: {"id": s["id"], "work_hours": s["work_hours"]} for s in shifts}

    schedules = db.execute(
        "SELECT employee_name, shift_name FROM raw_schedules WHERE schedule_date >= ? AND schedule_date <= ?",
        (start_date, end_date)
    ).fetchall()

    emp_shifts = {}
    for s in schedules:
        emp = s["employee_name"]
        shift = s["shift_name"]
        if emp not in emp_shifts:
            emp_shifts[emp] = {}
        emp_shifts[emp][shift] = emp_shifts[emp].get(shift, 0) + 1

    results = []
    for emp in employees:
        name = emp["name"]
        shifts_count = emp_shifts.get(name, {})

        total_hours = 0.0
        total_days = 0
        for shift_name, days in shifts_count.items():
            info = shift_map.get(shift_name)
            if info and info["work_hours"] > 0:
                total_hours += info["work_hours"] * days
                total_days += days

        work_system = (emp["work_hour_system"] or "").strip()
        if work_system == "综合计算工时制":
            system_hours = 167.0
        else:
            system_hours = 8.0 * count_working_days(year_int, month_int)

        diff = round(total_hours - system_hours, 1)

        results.append({
            "dongfu_id": emp["dongfu_id"] or "",
            "name": name,
            "team": emp["team"] or "",
            "work_system": work_system,
            "scheduled_hours": round(total_hours, 1),
            "system_hours": round(system_hours, 1),
            "diff": diff
        })

    # 返回所有可选团队（用于筛选器）
    all_teams = db.execute("SELECT DISTINCT team FROM employees WHERE status='active' ORDER BY team").fetchall()
    team_list = [r["team"] for r in all_teams]

    return jsonify({"rows": results, "teams": team_list})


@app.route("/api/monthly-hours-stats/export", methods=["GET"])
def api_monthly_hours_stats_export():
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    work_system = request.args.get("work_system", "").strip()
    teams_str = request.args.get("teams", "").strip()

    now = datetime.now()
    if not year:
        year = str(now.year)
    if not month:
        month = str(now.month)

    try:
        year_int = int(year)
        month_int = int(month)
    except (ValueError, TypeError):
        return jsonify({"error": "年份或月份格式无效"}), 400
    last_day = calendar.monthrange(year_int, month_int)[1]
    start_date = f"{year_int}-{month_int:02d}-01"
    end_date = f"{year_int}-{month_int:02d}-{last_day}"

    db = get_db()

    query = "SELECT id, name, team, dongfu_id, work_hour_system FROM employees WHERE status='active'"
    params = []
    if work_system:
        query += " AND work_hour_system = ?"
        params.append(work_system)
    teams = [t.strip() for t in teams_str.split(",") if t.strip()]
    if teams:
        placeholders = ",".join("?" for _ in teams)
        query += f" AND team IN ({placeholders})"
        params.extend(teams)
    query += " ORDER BY id"
    employees = db.execute(query, params).fetchall()

    shifts = db.execute("SELECT id, name, work_hours FROM shifts").fetchall()
    shift_map = {s["name"]: {"id": s["id"], "work_hours": s["work_hours"]} for s in shifts}

    schedules = db.execute(
        "SELECT employee_name, shift_name FROM raw_schedules WHERE schedule_date >= ? AND schedule_date <= ?",
        (start_date, end_date)
    ).fetchall()

    emp_shifts = {}
    for s in schedules:
        emp = s["employee_name"]
        shift = s["shift_name"]
        if emp not in emp_shifts:
            emp_shifts[emp] = {}
        emp_shifts[emp][shift] = emp_shifts[emp].get(shift, 0) + 1

    TEAM_ORDER = ["在线组","热线组","售后组","综合组","VIP组","质检组","支持组"]

    results = []
    for emp in employees:
        name = emp["name"]
        shifts_count = emp_shifts.get(name, {})

        total_hours = 0.0
        total_days = 0
        for shift_name, days in shifts_count.items():
            info = shift_map.get(shift_name)
            if info and info["work_hours"] > 0:
                total_hours += info["work_hours"] * days
                total_days += days

        ws = (emp["work_hour_system"] or "").strip()
        if ws == "综合计算工时制":
            system_hours = 167.0
        else:
            system_hours = 8.0 * count_working_days(year_int, month_int)

        diff = round(total_hours - system_hours, 1)

        results.append({
            "team": emp["team"] or "",
            "dongfu_id": emp["dongfu_id"] or "",
            "name": name,
            "scheduled_hours": round(total_hours, 1),
            "system_hours": round(system_hours, 1),
            "diff": diff
        })

    # 排序
    results.sort(key=lambda r: (TEAM_ORDER.index(r["team"]) if r["team"] in TEAM_ORDER else 999, r["name"]))

    # 生成 Excel
    wb = Workbook()
    ws_sheet = wb.active
    ws_sheet.title = f"{year_int}年{month_int}月时长统计"
    ws_sheet.append(["所属团队", "东福工号", "员工姓名", "原始排班时长(h)", "工时制度时长(h)", "差额(h)"])

    # 团队列颜色
    team_colors = {
        "在线组": "c6efce", "热线组": "b4d6fd", "售后组": "fcd5a7",
        "综合组": "e4c6ec", "VIP组": "f7c6d7", "质检组": "fce9a2", "支持组": "b9e4e0"
    }
    from openpyxl.styles import PatternFill, Font

    for row in results:
        ws_sheet.append([row["team"], row["dongfu_id"], row["name"],
                         row["scheduled_hours"], row["system_hours"], row["diff"]])
        r = ws_sheet.max_row
        color = team_colors.get(row["team"])
        if color:
            ws_sheet.cell(row=r, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # 差额颜色
        diff_cell = ws_sheet.cell(row=r, column=6)
        if row["diff"] > 0:
            diff_cell.font = Font(color="10b981", bold=True)
        elif row["diff"] < 0:
            diff_cell.font = Font(color="ef4444", bold=True)

    ws_sheet.column_dimensions["A"].width = 12
    ws_sheet.column_dimensions["B"].width = 14
    ws_sheet.column_dimensions["C"].width = 12
    ws_sheet.column_dimensions["D"].width = 18
    ws_sheet.column_dimensions["E"].width = 18
    ws_sheet.column_dimensions["F"].width = 12

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name=f"月度时长统计_{year_int}年{month_int}月.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 数据库管理 ====================

@app.route("/api/db/download")
def api_db_download():
    """下载数据库文件"""
    return send_file(DATABASE, as_attachment=True, download_name="schedule.db")


@app.route("/api/db/stats")
def api_db_stats():
    """数据库统计信息"""
    db = get_db()
    emp_count = db.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    shift_count = db.execute("SELECT COUNT(*) FROM shifts").fetchone()[0]
    sch_count = db.execute("SELECT COUNT(*) FROM schedules").fetchone()[0]
    date_range = db.execute("SELECT MIN(schedule_date), MAX(schedule_date) FROM schedules").fetchone()
    return jsonify({
        "employees": emp_count,
        "shifts": shift_count,
        "schedules": sch_count,
        "dateFrom": date_range[0] or "",
        "dateTo": date_range[1] or ""
    })

# ==================== 静态文件 ====================

@app.route("/")
def index():
    return app.send_static_file("排班表.html")


# ==================== 启动 ====================

if __name__ == "__main__":
    setup_logging()
    init_db()
    print("=" * 50)
    print("  客服排班系统后端已启动")
    print("  打开浏览器访问: http://127.0.0.1:5000")
    print("=" * 50)
    try:
        from waitress import serve
        print("  使用 waitress 生产服务器")
        serve(app, host="0.0.0.0", port=5000, threads=4)
    except ImportError:
        print("  waitress 未安装，使用开发服务器")
        print("  建议执行: pip install waitress")
        app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
